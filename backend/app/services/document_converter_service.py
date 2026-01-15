"""Document Converter Service for Legal AI Vault"""
from typing import Optional
import logging
import io
import tempfile
import os
import re
from io import BytesIO

logger = logging.getLogger(__name__)


class DocumentConverterService:
    """Service for converting documents to HTML format"""
    
    def __init__(self):
        """Initialize document converter service"""
        pass
    
    def convert_to_html(
        self,
        file_content: bytes,
        filename: str,
        file_type: Optional[str] = None
    ) -> str:
        """
        Convert document to HTML format
        
        Args:
            file_content: File content as bytes
            filename: Original filename
            file_type: File type/extension (optional, will be inferred from filename)
            
        Returns:
            HTML string representation of the document
            
        Raises:
            ValueError: If file format is not supported or conversion fails
        """
        if not file_content:
            raise ValueError("File content is empty")
        
        # Determine file type from filename if not provided
        if not file_type:
            _, ext = os.path.splitext(filename)
            file_type = ext.lower().lstrip('.')
        
        logger.info(f"Converting {filename} (type: {file_type}) to HTML")
        
        try:
            # Route to appropriate converter based on file type
            if file_type in ['docx', 'doc']:
                return self._convert_docx(file_content)
            elif file_type == 'pdf':
                return self._convert_pdf(file_content)
            elif file_type in ['xlsx', 'xls']:
                return self._convert_xlsx(file_content, filename)
            elif file_type == 'txt':
                return self._convert_txt(file_content)
            elif file_type in ['pptx', 'ppt']:
                return self._convert_with_markitdown(file_content, filename)
            else:
                # Try universal converter for other formats
                logger.warning(f"Unknown file type {file_type}, trying universal converter")
                return self._convert_with_markitdown(file_content, filename)
        except Exception as e:
            logger.error(f"Error converting {filename} to HTML: {e}", exc_info=True)
            raise ValueError(f"Не удалось конвертировать файл {filename}: {str(e)}")
    
    def _convert_docx(self, file_content: bytes) -> str:
        """
        Convert DOCX to HTML using mammoth
        
        Args:
            file_content: DOCX file content as bytes
            
        Returns:
            HTML string
        """
        try:
            import mammoth
            
            # Convert DOCX to HTML
            result = mammoth.convert_to_html(BytesIO(file_content))
            html = result.value
            
            # Log warnings if any
            if result.messages:
                logger.warning(f"Mammoth conversion warnings: {result.messages}")
            
            # Wrap in a container div for consistent styling
            html = f'<div class="docx-content">{html}</div>'
            
            logger.info("Successfully converted DOCX to HTML")
            return html
        except ImportError:
            logger.error("mammoth library not installed")
            raise ValueError("Библиотека для конвертации DOCX не установлена")
        except Exception as e:
            logger.error(f"Error converting DOCX: {e}", exc_info=True)
            raise ValueError(f"Ошибка конвертации DOCX: {str(e)}")
    
    def _convert_pdf(self, file_content: bytes) -> str:
        """
        Convert PDF to HTML using pypdf
        
        Args:
            file_content: PDF file content as bytes
            
        Returns:
            HTML string
        """
        try:
            from pypdf import PdfReader
            
            # Read PDF
            pdf_reader = PdfReader(BytesIO(file_content))
            
            # Extract text from all pages
            html_parts = ['<div class="pdf-content">']
            
            for page_num, page in enumerate(pdf_reader.pages, 1):
                try:
                    text = page.extract_text()
                    if text:
                        # Split text into paragraphs (double newlines)
                        paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]
                        
                        html_parts.append(f'<div class="pdf-page" data-page="{page_num}">')
                        for para in paragraphs:
                            # Replace single newlines with <br> within paragraphs
                            para_html = para.replace('\n', '<br/>')
                            html_parts.append(f'<p>{para_html}</p>')
                        html_parts.append('</div>')
                except Exception as e:
                    logger.warning(f"Error extracting text from page {page_num}: {e}")
                    html_parts.append(f'<div class="pdf-page" data-page="{page_num}"><p class="error">Не удалось извлечь текст со страницы {page_num}</p></div>')
            
            html_parts.append('</div>')
            
            html = '\n'.join(html_parts)
            
            if not html or html == '<div class="pdf-content"></div>':
                raise ValueError("Не удалось извлечь текст из PDF файла")
            
            logger.info(f"Successfully converted PDF to HTML ({len(pdf_reader.pages)} pages)")
            return html
        except ImportError:
            logger.error("pypdf library not installed")
            raise ValueError("Библиотека для конвертации PDF не установлена")
        except Exception as e:
            logger.error(f"Error converting PDF: {e}", exc_info=True)
            raise ValueError(f"Ошибка конвертации PDF: {str(e)}")
    
    def _convert_xlsx(self, file_content: bytes, filename: str) -> str:
        """
        Convert XLSX/XLS to HTML using openpyxl
        
        Args:
            file_content: Excel file content as bytes
            filename: Original filename
            
        Returns:
            HTML string with tables
        """
        try:
            import openpyxl
            
            # Load workbook
            workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
            
            html_parts = ['<div class="xlsx-content">']
            
            # Convert each sheet to HTML table
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                html_parts.append(f'<div class="xlsx-sheet">')
                html_parts.append(f'<h3 class="sheet-name">{sheet_name}</h3>')
                html_parts.append('<table class="xlsx-table">')
                
                # Get dimensions
                if sheet.max_row == 0 or sheet.max_column == 0:
                    html_parts.append('<tr><td colspan="1">Пустой лист</td></tr>')
                else:
                    # Add header row if first row looks like headers
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                        html_parts.append('<tr>')
                        for cell_value in row:
                            cell_html = str(cell_value) if cell_value is not None else ''
                            # Escape HTML special characters
                            cell_html = cell_html.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                            
                            # Use th for first row if it looks like headers
                            tag = 'th' if row_idx == 1 else 'td'
                            html_parts.append(f'<{tag}>{cell_html}</{tag}>')
                        html_parts.append('</tr>')
                
                html_parts.append('</table>')
                html_parts.append('</div>')
            
            html_parts.append('</div>')
            
            html = '\n'.join(html_parts)
            
            logger.info(f"Successfully converted XLSX to HTML ({len(workbook.sheetnames)} sheets)")
            return html
        except ImportError:
            logger.error("openpyxl library not installed")
            raise ValueError("Библиотека для конвертации Excel не установлена")
        except Exception as e:
            logger.error(f"Error converting XLSX: {e}", exc_info=True)
            raise ValueError(f"Ошибка конвертации Excel: {str(e)}")
    
    def _convert_txt(self, file_content: bytes) -> str:
        """
        Convert TXT to HTML
        
        Args:
            file_content: Text file content as bytes
            
        Returns:
            HTML string
        """
        try:
            # Try UTF-8 first, fallback to other encodings
            try:
                text = file_content.decode('utf-8')
            except UnicodeDecodeError:
                try:
                    text = file_content.decode('windows-1251')
                except UnicodeDecodeError:
                    text = file_content.decode('latin-1', errors='ignore')
            
            # Split into paragraphs (double newlines)
            paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]
            
            if not paragraphs:
                # If no double newlines, split by single newlines
                paragraphs = [p.strip() for p in text.split('\n') if p.strip()]
            
            html_parts = ['<div class="txt-content">']
            
            for para in paragraphs:
                # Replace single newlines with <br> within paragraphs
                para_html = para.replace('\n', '<br/>')
                # Escape HTML special characters
                para_html = para_html.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                html_parts.append(f'<p>{para_html}</p>')
            
            html_parts.append('</div>')
            
            html = '\n'.join(html_parts)
            
            logger.info("Successfully converted TXT to HTML")
            return html
        except Exception as e:
            logger.error(f"Error converting TXT: {e}", exc_info=True)
            raise ValueError(f"Ошибка конвертации TXT: {str(e)}")
    
    def _convert_with_markitdown(
        self,
        file_content: bytes,
        filename: str
    ) -> str:
        """
        Convert document to HTML using markitdown (universal converter)
        
        Args:
            file_content: File content as bytes
            filename: Original filename
            
        Returns:
            HTML string
        """
        try:
            from markitdown import MarkItDown
            import markdown
            
            # Save to temporary file for markitdown
            _, ext = os.path.splitext(filename)
            with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp_file:
                tmp_file.write(file_content)
                tmp_path = tmp_file.name
            
            try:
                # Convert using markitdown
                md = MarkItDown()
                result = md.convert(tmp_path)
                
                # Get markdown content
                markdown_content = result.text_content
                
                # Convert markdown to HTML
                html = markdown.markdown(markdown_content, extensions=['extra', 'nl2br'])
                
                # Wrap in container
                html = f'<div class="markitdown-content">{html}</div>'
                
                logger.info(f"Successfully converted {filename} to HTML using markitdown")
                return html
            finally:
                # Clean up temp file
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
        except ImportError:
            logger.warning("markitdown or markdown library not installed, falling back to basic conversion")
            # Fallback: try to extract text and wrap in HTML
            try:
                text = file_content.decode('utf-8', errors='ignore')
                html = f'<div class="fallback-content"><pre>{text}</pre></div>'
                return html
            except Exception as e:
                logger.error(f"Fallback conversion failed: {e}", exc_info=True)
                raise ValueError(f"Не удалось конвертировать файл {filename}: формат не поддерживается")
        except Exception as e:
            logger.error(f"Error converting with markitdown: {e}", exc_info=True)
            # Fallback to basic text extraction
            try:
                text = file_content.decode('utf-8', errors='ignore')
                html = f'<div class="fallback-content"><pre>{text}</pre></div>'
                logger.warning(f"Used fallback conversion for {filename}")
                return html
            except Exception as fallback_error:
                logger.error(f"Fallback conversion also failed: {fallback_error}", exc_info=True)
                raise ValueError(f"Не удалось конвертировать файл {filename}: {str(e)}")

