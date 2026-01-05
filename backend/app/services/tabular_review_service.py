"""Tabular Review service for Legal AI Vault"""
from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_
from app.models.tabular_review import (
    TabularReview, TabularColumn, TabularCell, 
    TabularColumnTemplate, TabularDocumentStatus
)
from app.models.case import Case, File
from app.models.user import User
from app.services.llm_factory import create_llm
from app.config import config
from app.services.tabular_review_models import TabularCellExtractionModel
from langchain_core.documents import Document
from langchain_core.prompts import ChatPromptTemplate
import logging
import asyncio
from datetime import datetime, timedelta

logger = logging.getLogger(__name__)


class TabularReviewService:
    """Service for managing tabular reviews"""
    
    def __init__(self, db: Session):
        """Initialize tabular review service"""
        self.db = db
        # Initialize LLM for extraction
        try:
            self.llm = create_llm(temperature=0.1)  # Low temperature for deterministic extraction
        except Exception as e:
            self.llm = None
            logger.warning(f"GigaChat not configured: {e}, extraction will not work")
    
    def create_tabular_review(
        self, 
        case_id: str, 
        user_id: str,
        name: str, 
        description: Optional[str] = None,
        selected_file_ids: Optional[List[str]] = None
    ) -> TabularReview:
        """Create a new tabular review"""
        # Verify case exists and belongs to user
        case = self.db.query(Case).filter(
            and_(Case.id == case_id, Case.user_id == user_id)
        ).first()
        
        if not case:
            raise ValueError(f"Case {case_id} not found or access denied")
        
        # Verify selected files belong to the case
        if selected_file_ids:
            files = self.db.query(File).filter(
                and_(
                    File.id.in_(selected_file_ids),
                    File.case_id == case_id
                )
            ).all()
            if len(files) != len(selected_file_ids):
                raise ValueError("Some selected files do not belong to this case")
        
        review = TabularReview(
            case_id=case_id,
            user_id=user_id,
            name=name,
            description=description,
            status="draft",
            selected_file_ids=selected_file_ids
        )
        
        self.db.add(review)
        self.db.commit()
        self.db.refresh(review)
        
        logger.info(f"Created tabular review {review.id} for case {case_id} with {len(selected_file_ids) if selected_file_ids else 0} selected files")
        return review
    
    def add_column(
        self,
        review_id: str,
        column_label: str,
        column_type: str,
        prompt: str,
        user_id: str,
        column_config: Optional[Dict[str, Any]] = None
    ) -> TabularColumn:
        """Add a column to tabular review"""
        # Verify review belongs to user
        review = self.db.query(TabularReview).filter(
            and_(TabularReview.id == review_id, TabularReview.user_id == user_id)
        ).first()
        
        if not review:
            raise ValueError(f"Tabular review {review_id} not found or access denied")
        
        # Get max order_index
        max_order = self.db.query(TabularColumn.order_index).filter(
            TabularColumn.tabular_review_id == review_id
        ).order_by(TabularColumn.order_index.desc()).first()
        
        order_index = (max_order[0] + 1) if max_order else 0
        
        column = TabularColumn(
            tabular_review_id=review_id,
            column_label=column_label,
            column_type=column_type,
            prompt=prompt,
            column_config=column_config,
            order_index=order_index
        )
        
        self.db.add(column)
        self.db.commit()
        self.db.refresh(column)
        
        logger.info(f"Added column {column.id} to review {review_id}")
        return column
    
    def update_column(
        self,
        review_id: str,
        column_id: str,
        user_id: str,
        column_label: Optional[str] = None,
        prompt: Optional[str] = None,
        column_config: Optional[Dict[str, Any]] = None
    ) -> TabularColumn:
        """Update a column (rename, update prompt, config)"""
        # Verify review belongs to user
        review = self.db.query(TabularReview).filter(
            and_(TabularReview.id == review_id, TabularReview.user_id == user_id)
        ).first()
        
        if not review:
            raise ValueError(f"Tabular review {review_id} not found or access denied")
        
        # Get the column
        column = self.db.query(TabularColumn).filter(
            and_(
                TabularColumn.id == column_id,
                TabularColumn.tabular_review_id == review_id
            )
        ).first()
        
        if not column:
            raise ValueError(f"Column {column_id} not found in review {review_id}")
        
        # Update fields
        if column_label is not None:
            column.column_label = column_label
        if prompt is not None:
            column.prompt = prompt
        if column_config is not None:
            column.column_config = column_config
        
        self.db.commit()
        self.db.refresh(column)
        
        logger.info(f"Updated column {column_id} in review {review_id}")
        return column
    
    def delete_column(
        self,
        review_id: str,
        column_id: str,
        user_id: str
    ) -> bool:
        """Delete a column (cascade deletes all cells)"""
        # Verify review belongs to user
        review = self.db.query(TabularReview).filter(
            and_(TabularReview.id == review_id, TabularReview.user_id == user_id)
        ).first()
        
        if not review:
            raise ValueError(f"Tabular review {review_id} not found or access denied")
        
        # Get the column
        column = self.db.query(TabularColumn).filter(
            and_(
                TabularColumn.id == column_id,
                TabularColumn.tabular_review_id == review_id
            )
        ).first()
        
        if not column:
            raise ValueError(f"Column {column_id} not found in review {review_id}")
        
        # Delete column (cascade will delete all cells)
        self.db.delete(column)
        self.db.commit()
        
        logger.info(f"Deleted column {column_id} from review {review_id}")
        return True
    
    def reorder_columns(
        self,
        review_id: str,
        column_ids: List[str],
        user_id: str
    ) -> List[TabularColumn]:
        """Reorder columns by providing ordered list of column IDs"""
        # Verify review belongs to user
        review = self.db.query(TabularReview).filter(
            and_(TabularReview.id == review_id, TabularReview.user_id == user_id)
        ).first()
        
        if not review:
            raise ValueError(f"Tabular review {review_id} not found or access denied")
        
        # Get all columns for this review
        columns = self.db.query(TabularColumn).filter(
            TabularColumn.tabular_review_id == review_id
        ).all()
        
        column_map = {col.id: col for col in columns}
        
        # Verify all provided IDs exist and belong to this review
        for col_id in column_ids:
            if col_id not in column_map:
                raise ValueError(f"Column {col_id} not found in review {review_id}")
        
        # Update order_index for each column
        for index, col_id in enumerate(column_ids):
            column_map[col_id].order_index = index
        
        # Handle columns not in the list (shouldn't happen, but be safe)
        for col in columns:
            if col.id not in column_ids:
                # Put them at the end
                col.order_index = len(column_ids) + columns.index(col)
        
        self.db.commit()
        
        # Return columns in new order
        updated_columns = self.db.query(TabularColumn).filter(
            TabularColumn.tabular_review_id == review_id
        ).order_by(TabularColumn.order_index).all()
        
        logger.info(f"Reordered columns for review {review_id}")
        return updated_columns
    
    def get_table_data(self, review_id: str, user_id: str) -> Dict[str, Any]:
        """Get table data for tabular review"""
        # Verify review belongs to user
        review = self.db.query(TabularReview).filter(
            and_(TabularReview.id == review_id, TabularReview.user_id == user_id)
        ).first()
        
        if not review:
            raise ValueError(f"Tabular review {review_id} not found or access denied")
        
        # Get files for the case, filtered by selected_file_ids if specified
        files_query = self.db.query(File).filter(File.case_id == review.case_id)
        if review.selected_file_ids:
            files_query = files_query.filter(File.id.in_(review.selected_file_ids))
        files = files_query.all()
        
        # Get columns
        columns = self.db.query(TabularColumn).filter(
            TabularColumn.tabular_review_id == review_id
        ).order_by(TabularColumn.order_index).all()
        
        # Get all cells
        cells = self.db.query(TabularCell).filter(
            TabularCell.tabular_review_id == review_id
        ).all()
        
        # Get document statuses
        statuses = self.db.query(TabularDocumentStatus).filter(
            TabularDocumentStatus.tabular_review_id == review_id,
            TabularDocumentStatus.user_id == user_id
        ).all()
        
        # Build cell map for quick lookup
        cell_map = {}
        for cell in cells:
            key = (cell.file_id, cell.column_id)
            cell_map[key] = cell
        
        # Build status map
        status_map = {s.file_id: s for s in statuses}
        
        # Build table data structure
        table_rows = []
        for file in files:
            row = {
                "file_id": file.id,
                "file_name": file.filename,
                "file_type": file.file_type,
                "status": status_map.get(file.id, {}).status if file.id in status_map else "not_reviewed",
                "cells": {}
            }
            
            for column in columns:
                key = (file.id, column.id)
                cell = cell_map.get(key)
                cell_value = cell.cell_value if cell else None
                
                # Parse JSON if cell_value is a JSON string
                if cell_value and isinstance(cell_value, str) and cell_value.strip().startswith("{"):
                    try:
                        import json
                        parsed = json.loads(cell_value)
                        if isinstance(parsed, dict) and "cell_value" in parsed:
                            cell_value = parsed["cell_value"]
                    except (json.JSONDecodeError, ValueError):
                        # Not valid JSON, keep original value
                        pass
                
                row["cells"][column.id] = {
                    "cell_value": cell_value,
                    "verbatim_extract": cell.verbatim_extract if cell else None,
                    "reasoning": cell.reasoning if cell else None,
                    "confidence_score": float(cell.confidence_score) if cell and cell.confidence_score else None,
                    "source_page": cell.source_page if cell else None,
                    "source_section": cell.source_section if cell else None,
                    "status": cell.status if cell else "pending",
                }
            
            table_rows.append(row)
        
        return {
            "review": {
                "id": review.id,
                "name": review.name,
                "description": review.description,
                "status": review.status,
                "selected_file_ids": review.selected_file_ids,
            },
            "columns": [
                {
                    "id": col.id,
                    "column_label": col.column_label,
                    "column_type": col.column_type,
                    "prompt": col.prompt,
                    "order_index": col.order_index,
                }
                for col in columns
            ],
            "rows": table_rows,
        }
    
    async def extract_cell_value(
        self,
        file: File,
        column: TabularColumn
    ) -> Dict[str, Any]:
        """Extract cell value for a specific file and column"""
        try:
            # Get document text
            document_text = file.original_text or ""
            if not document_text:
                logger.warning(f"File {file.id} has no text content")
                return {
                    "file_id": file.id,
                    "column_id": column.id,
                    "cell_value": None,
                    "error": "No text content"
                }
            
            # Limit text to avoid token limits (use first 8000 chars)
            limited_text = document_text[:8000]
            
            # Build prompt based on column type
            column_type_descriptions = {
                "text": "свободный текст",
                "bulleted_list": "маркированный список (каждый пункт с новой строки, начинается с •)",
                "number": "числовое значение",
                "currency": "денежная сумма с валютой (например: '100000 USD' или '50 000 руб.')",
                "yes_no": "только 'Yes' или 'No'",
                "date": "дата в формате YYYY-MM-DD или DD.MM.YYYY",
                "tag": "один тег из предопределенного списка",
                "multiple_tags": "несколько тегов из предопределенного списка (через запятую)",
                "verbatim": "точная цитата из документа с указанием источника",
                "manual_input": "ручной ввод (не используется AI)"
            }
            
            type_desc = column_type_descriptions.get(column.column_type, "свободный текст")
            
            # Для tag/multiple_tags добавляем доступные опции
            tag_options_text = ""
            if column.column_type in ["tag", "multiple_tags"] and column.column_config:
                options = column.column_config.get("options", [])
                if options:
                    tag_options_text = f"\n\nДоступные опции: {', '.join([opt.get('label', '') for opt in options])}"
            
            system_prompt = f"""Ты эксперт по извлечению информации из юридических документов.
Твоя задача - ответить на вопрос о документе и предоставить подробное обоснование.

Тип ответа: {column.column_type}
Описание: {type_desc}{tag_options_text}

ВАЖНО:
1. Если информация не найдена, верни "N/A" для cell_value
2. Для yes_no: только "Yes" или "No" или "Unknown"
3. Для verbatim: приведи точную цитату из документа
4. Для tag/multiple_tags: используй ТОЛЬКО опции из списка выше
5. ВСЕГДА указывай reasoning - подробное объяснение, откуда взялась информация
6. ВСЕГДА указывай source_references - конкретные места в документе (страницы, разделы, цитаты)

Формат ответа (JSON):
{{
    "cell_value": "извлеченное значение",
    "reasoning": "подробное объяснение, почему именно такой ответ, с указанием конкретных мест в документе",
    "source_references": [
        {{"page": 1, "section": "Раздел 3.1", "text": "цитата из документа"}},
        {{"page": 2, "section": null, "text": "еще одна цитата"}}
    ],
    "confidence": 0.9
}}"""

            user_prompt = f"""Вопрос: {column.prompt}

Документ ({file.filename}):
{limited_text}

Извлеки информацию согласно типу {column.column_type} и верни JSON с cell_value, reasoning и source_references."""
            
            if not self.llm:
                raise ValueError("LLM not configured")
            
            # Try structured output first, fallback to regular if not supported
            try:
                from langchain_core.pydantic_v1 import BaseModel as PydanticBaseModel
                from typing import List as TypingList
                
                class ExtractionResponse(PydanticBaseModel):
                    cell_value: str
                    reasoning: str
                    source_references: TypingList[dict] = []
                    confidence: float = 0.85
                
                # Use structured output if LLM supports it
                if hasattr(self.llm, 'with_structured_output'):
                    structured_llm = self.llm.with_structured_output(ExtractionResponse)
                    from langchain_core.messages import SystemMessage, HumanMessage
                    messages = [
                        SystemMessage(content=system_prompt),
                        HumanMessage(content=user_prompt)
                    ]
                    result = await structured_llm.ainvoke(messages)
                    
                    cell_value = result.cell_value
                    reasoning = result.reasoning
                    source_references = result.source_references or []
                    confidence = result.confidence
                else:
                    # Fallback to regular LLM call with JSON parsing
                    from langchain_core.messages import SystemMessage, HumanMessage
                    messages = [
                        SystemMessage(content=system_prompt),
                        HumanMessage(content=user_prompt)
                    ]
                    response = await self.llm.ainvoke(messages)
                    response_text = response.content.strip() if hasattr(response, 'content') else str(response).strip()
                    
                    # Filter out refusal responses from LLM
                    refusal_patterns = [
                        r'не обладаю.*мнением',
                        r'не транслирую.*мнение',
                        r'ответ сгенерирован.*моделью',
                        r'разговоры.*ограничены',
                        r'не могу.*ответить',
                        r'ограничены.*темы',
                        r'неправильного толкования',
                        r'чувствительные темы',
                    ]
                    for pattern in refusal_patterns:
                        if re.search(pattern, response_text, re.IGNORECASE):
                            logger.warning(f"LLM returned refusal response, retrying with different prompt")
                            # Retry with a more direct prompt
                            retry_prompt = f"Извлеки информацию из документа. Вопрос: {column.prompt}\n\nДокумент:\n{limited_text}\n\nВерни только JSON с полями cell_value, reasoning, source_references."
                            retry_messages = [
                                SystemMessage(content="Ты помощник для извлечения информации из юридических документов. Всегда отвечай на русском языке."),
                                HumanMessage(content=retry_prompt)
                            ]
                            response = await self.llm.ainvoke(retry_messages)
                            response_text = response.content.strip() if hasattr(response, 'content') else str(response).strip()
                            break
                    
                    # Try to parse JSON from response
                    import json
                    import re
                    # Extract JSON from response (might be wrapped in markdown code blocks)
                    # Improved regex to match nested JSON objects
                    json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*"cell_value"[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', response_text, re.DOTALL)
                    if json_match:
                        try:
                            parsed = json.loads(json_match.group(0))
                            cell_value = parsed.get("cell_value", response_text)
                            reasoning = parsed.get("reasoning", f"Извлечено из документа '{file.filename}'")
                            source_references = parsed.get("source_references", [])
                            confidence = parsed.get("confidence", 0.85)
                        except:
                            cell_value = response_text
                            reasoning = f"Извлечено из документа '{file.filename}' на основе вопроса: {column.prompt}"
                            source_references = []
                            confidence = 0.85
                    else:
                        cell_value = response_text
                        reasoning = f"Извлечено из документа '{file.filename}' на основе вопроса: {column.prompt}"
                        source_references = []
                        confidence = 0.85
            except Exception as e:
                logger.warning(f"Structured output failed, using fallback: {e}")
                # Fallback to simple extraction
                from langchain_core.messages import SystemMessage, HumanMessage
                messages = [
                    SystemMessage(content=system_prompt),
                    HumanMessage(content=user_prompt)
                ]
                response = await self.llm.ainvoke(messages)
                cell_value = response.content.strip() if hasattr(response, 'content') else str(response).strip()
                reasoning = f"Извлечено из документа '{file.filename}' на основе вопроса: {column.prompt}"
                source_references = []
                confidence = 0.85
            
            # Format value based on type
            if column.column_type == "yes_no":
                if cell_value.lower() in ["yes", "да", "есть", "true"]:
                    cell_value = "Yes"
                elif cell_value.lower() in ["no", "нет", "нету", "false"]:
                    cell_value = "No"
                else:
                    cell_value = "Unknown"
            elif column.column_type == "bulleted_list":
                # Ensure bulleted list format
                if cell_value and not cell_value.startswith("•"):
                    lines = [line.strip() for line in cell_value.split("\n") if line.strip()]
                    cell_value = "\n".join([f"• {line}" if not line.startswith("•") else line for line in lines])
            elif column.column_type in ["tag", "multiple_tags"]:
                # Validate against column_config options
                if column.column_config and column.column_config.get("options"):
                    valid_options = [opt.get("label", "").lower() for opt in column.column_config["options"]]
                    # Split by comma for multiple_tags
                    if column.column_type == "multiple_tags":
                        tags = [tag.strip() for tag in cell_value.split(",")]
                        # Filter to only valid options
                        valid_tags = [tag for tag in tags if tag.lower() in valid_options]
                        cell_value = ", ".join(valid_tags) if valid_tags else cell_value
                    else:
                        # Single tag - check if valid
                        if cell_value.lower() not in valid_options:
                            # Try to find closest match
                            for opt in column.column_config["options"]:
                                if opt.get("label", "").lower() in cell_value.lower():
                                    cell_value = opt.get("label", "")
                                    break
            
            # Extract verbatim if type is verbatim
            verbatim_extract = cell_value if column.column_type == "verbatim" else None
            
            return {
                "file_id": file.id,
                "column_id": column.id,
                "cell_value": cell_value,
                "verbatim_extract": verbatim_extract,
                "reasoning": reasoning,
                "source_references": source_references,
                "confidence_score": confidence,
            }
            
        except Exception as e:
            logger.error(f"Error extracting cell value for file {file.id}, column {column.id}: {e}", exc_info=True)
            return {
                "file_id": file.id,
                "column_id": column.id,
                "cell_value": None,
                "error": str(e)
            }
    
    def _is_complex_question(self, prompt: str) -> bool:
        """Determine if question is complex and requires advanced reasoning"""
        complex_keywords = [
            "оцени", "проанализируй", "сравни", "найди все",
            "риски", "противоречия", "последствия", "прецеденты",
            "юридические", "правовые", "судебная практика", "анализ"
        ]
        prompt_lower = prompt.lower()
        return any(keyword in prompt_lower for keyword in complex_keywords)
    
    def _determine_extraction_strategy(
        self,
        file: File,
        column: TabularColumn
    ) -> str:
        """
        Determine optimal extraction strategy based on column type, question complexity, and document size
        
        Returns:
            Strategy name: 'simple', 'multi_query', 'compression', 'iterative', 'map_reduce'
        """
        doc_length = len(file.original_text or "")
        is_complex = self._is_complex_question(column.prompt)
        
        # Simple types (date, currency, number, yes_no) - use simpler approach
        if column.column_type in ['date', 'currency', 'number', 'yes_no']:
            if doc_length < 10000:
                return 'multi_query'  # Short document - use multi-query retriever
            else:
                return 'compression'  # Long document - use compression retriever
        
        # Complex types (text, verbatim) - use advanced approach
        elif column.column_type in ['text', 'verbatim']:
            if is_complex:
                return 'iterative'  # Complex question - use iterative RAG
            elif doc_length > 50000:
                return 'compression'  # Very long document - use compression
            else:
                return 'multi_query'  # Medium case - use multi-query
        
        # Default - multi-query
        return 'multi_query'
    
    async def extract_cell_value_improved(
        self,
        file: File,
        column: TabularColumn,
        strategy: str = 'multi_query'
    ) -> Dict[str, Any]:
        """
        Improved extraction using LangChain retrievers and structured output
        
        Args:
            file: File to extract from
            column: Column definition
            strategy: Extraction strategy ('multi_query', 'compression', 'iterative')
        
        Returns:
            Dictionary with extraction results
        """
        try:
            from app.services.document_processor import DocumentProcessor
            from app.services.langchain_retrievers import AdvancedRetrieverService
            from app.services.iterative_rag_service import IterativeRAGService
            from app.services.rag_service import RAGService
            from app.services.legal_splitter import LegalTextSplitter
            from app.services.langchain_agents.llm_helper import create_fixing_parser
            
            document_processor = DocumentProcessor()
            retriever_service = AdvancedRetrieverService(document_processor)
            
            # Get document text
            document_text = file.original_text or ""
            if not document_text:
                logger.warning(f"File {file.id} has no text content")
                return {
                    "file_id": file.id,
                    "column_id": column.id,
                    "cell_value": None,
                    "error": "No text content"
                }
            
            # Build query for retrieval
            query = f"{column.prompt} {column.column_label}"
            
            # Retrieve relevant chunks based on strategy
            relevant_chunks = []
            
            try:
                if strategy == 'iterative':
                    # Use IterativeRAGService for complex questions
                    rag_service = RAGService()
                    iterative_rag = IterativeRAGService(rag_service)
                    relevant_chunks = iterative_rag.retrieve_iteratively(
                        case_id=file.case_id,
                        query=query,
                        max_iterations=3,
                        initial_k=5
                    )
                elif strategy == 'compression':
                    # Use compression retriever for long documents
                    relevant_chunks = retriever_service.retrieve_with_compression(
                        case_id=file.case_id,
                        query=query,
                        k=5
                    )
                else:
                    # Default: multi-query retriever
                    relevant_chunks = retriever_service.retrieve_with_multi_query(
                        case_id=file.case_id,
                        query=query,
                        k=5
                    )
            except Exception as e:
                logger.warning(f"Error retrieving chunks with strategy '{strategy}': {e}, falling back to document chunking")
                relevant_chunks = []
            
            # Filter chunks for this specific file
            file_chunks = [
                chunk for chunk in relevant_chunks
                if (chunk.metadata.get('file_id') == file.id or
                    chunk.metadata.get('source_file') == file.filename or
                    chunk.metadata.get('source_file', '').endswith(file.filename))
            ]
            
            # If no chunks found through retriever, use document chunking as fallback
            if not file_chunks:
                logger.debug(f"No chunks found via retriever for file {file.id}, using document chunking")
                splitter = LegalTextSplitter(chunk_size=1200, chunk_overlap=300)
                all_chunks = splitter.split_text(document_text)
                file_chunks = [
                    Document(
                        page_content=chunk,
                        metadata={
                            "chunk_index": i,
                            "file_id": file.id,
                            "source_file": file.filename
                        }
                    )
                    for i, chunk in enumerate(all_chunks)
                ]
            
            if not file_chunks:
                return {
                    "file_id": file.id,
                    "column_id": column.id,
                    "cell_value": None,
                    "error": "No chunks available for extraction"
                }
            
            # Try to use structured output
            best_result = None
            best_confidence = 0.0
            
            # Process top 3 chunks
            for chunk in file_chunks[:3]:
                try:
                    # Build type descriptions
                    column_type_descriptions = {
                        "text": "свободный текст",
                        "bulleted_list": "маркированный список (каждый пункт с новой строки, начинается с •)",
                        "number": "числовое значение",
                        "currency": "денежная сумма с валютой (например: '100000 USD' или '50 000 руб.')",
                        "yes_no": "только 'Yes' или 'No'",
                        "date": "дата в формате YYYY-MM-DD или DD.MM.YYYY",
                        "tag": "один тег из предопределенного списка",
                        "multiple_tags": "несколько тегов из предопределенного списка (через запятую)",
                        "verbatim": "точная цитата из документа с указанием источника",
                        "manual_input": "ручной ввод (не используется AI)"
                    }
                    
                    type_desc = column_type_descriptions.get(column.column_type, "свободный текст")
                    
                    # Для tag/multiple_tags добавляем доступные опции
                    tag_options_text = ""
                    if column.column_type in ["tag", "multiple_tags"] and column.column_config:
                        options = column.column_config.get("options", [])
                        if options:
                            tag_options_text = f"\n\nДоступные опции: {', '.join([opt.get('label', '') for opt in options])}"
                    
                    # Create prompt
                    prompt = ChatPromptTemplate.from_messages([
                        ("system", f"""Ты эксперт по извлечению информации из юридических документов.
Тип ответа: {column.column_type}
Описание: {type_desc}{tag_options_text}

ВАЖНО:
- Указывай source_page если возможно
- Указывай source_section (например, "Раздел 3.1", "Статья 5")
- Указывай source_references - список конкретных мест в документе с цитатами: [{{"page": 1, "section": "Раздел 3", "text": "цитата"}}]
- Указывай reasoning - подробное объяснение, почему именно такой ответ, с указанием конкретных мест в документе
- Указывай confidence - насколько ты уверен (0.0-1.0)
- Если информация не найдена, верни "N/A" для cell_value
- Для tag/multiple_tags используй ТОЛЬКО опции из списка выше"""),
                        ("human", f"""Вопрос: {column.prompt}

Документ:
{chunk.page_content}

Извлеки информацию согласно типу {column.column_type}.
Если это verbatim, приведи ТОЧНУЮ цитату из документа.
ВСЕГДА указывай source_references с конкретными цитатами из документа.""")
                    ])
                    
                    # Try structured output first
                    try:
                        structured_llm = self.llm.with_structured_output(
                            TabularCellExtractionModel,
                            method="json_schema"
                        )
                        chain = prompt | structured_llm
                        result = await chain.ainvoke({})
                        
                        # Add column_type for validation
                        result.column_type = column.column_type
                        
                    except Exception as e:
                        logger.warning(f"Structured output failed, using parser: {e}")
                        # Fallback: use parser with retry
                        chain = prompt | self.llm
                        response = await chain.ainvoke({})
                        response_text = response.content if hasattr(response, 'content') else str(response)
                        
                        parser = create_fixing_parser(
                            TabularCellExtractionModel,
                            self.llm,
                            max_retries=3
                        )
                        result = parser.parse(response_text)
                        result.column_type = column.column_type
                    
                    # Update source_page from metadata if not set
                    if not result.source_page:
                        result.source_page = chunk.metadata.get('source_page')
                    
                    # Update source_section if available in metadata
                    if not result.source_section:
                        result.source_section = chunk.metadata.get('source_section')
                    
                    # Validate result
                    if result.confidence > best_confidence:
                        best_result = result
                        best_confidence = result.confidence
                        
                        # Early stop if confidence is very high
                        if result.confidence >= 0.95:
                            break
                            
                except Exception as e:
                    logger.warning(f"Error extracting from chunk: {e}")
                    continue
            
            if not best_result:
                # Final fallback: use old method
                logger.warning(f"All extraction attempts failed for file {file.id}, column {column.id}, using fallback")
                return await self.extract_cell_value(file, column)
            
            # Build source_references from result
            source_references = []
            if best_result.source_page or best_result.source_section:
                source_ref = {
                    "page": best_result.source_page,
                    "section": best_result.source_section,
                    "text": best_result.verbatim_extract or best_result.cell_value[:200]  # First 200 chars as quote
                }
                source_references.append(source_ref)
            
            # Add any additional source_references from result
            if hasattr(best_result, 'source_references') and best_result.source_references:
                source_references.extend(best_result.source_references)
            
            return {
                "file_id": file.id,
                "column_id": column.id,
                "cell_value": best_result.cell_value,
                "verbatim_extract": best_result.verbatim_extract,
                "source_page": best_result.source_page,
                "source_section": best_result.source_section,
                "source_references": source_references,
                "reasoning": best_result.reasoning,
                "confidence_score": float(best_result.confidence),
                "extraction_method": best_result.extraction_method
            }
            
        except Exception as e:
            logger.error(f"Error in improved extraction for file {file.id}, column {column.id}: {e}", exc_info=True)
            # Fallback to old method
            return await self.extract_cell_value(file, column)
    
    async def extract_cell_value_smart(
        self,
        file: File,
        column: TabularColumn
    ) -> Dict[str, Any]:
        """
        Smart extraction with automatic strategy selection
        
        Determines optimal strategy and uses it for extraction
        """
        strategy = self._determine_extraction_strategy(file, column)
        logger.debug(f"Using extraction strategy '{strategy}' for file {file.id}, column {column.id}")
        return await self.extract_cell_value_improved(file, column, strategy=strategy)
    
    async def run_extraction(self, review_id: str, user_id: str) -> Dict[str, Any]:
        """Run parallel extraction for all documents and columns"""
        # Verify review belongs to user
        review = self.db.query(TabularReview).filter(
            and_(TabularReview.id == review_id, TabularReview.user_id == user_id)
        ).first()
        
        if not review:
            raise ValueError(f"Tabular review {review_id} not found or access denied")
        
        # Update status
        review.status = "processing"
        self.db.commit()
        
        try:
            # Get files - filter by selected_file_ids if specified
            files_query = self.db.query(File).filter(File.case_id == review.case_id)
            if review.selected_file_ids:
                files_query = files_query.filter(File.id.in_(review.selected_file_ids))
            files = files_query.all()
            
            # Get columns
            columns = self.db.query(TabularColumn).filter(
                TabularColumn.tabular_review_id == review_id
            ).order_by(TabularColumn.order_index).all()
            
            if not files:
                raise ValueError("No files found for this case")
            
            if not columns:
                raise ValueError("No columns defined for this review")
            
            # Create tasks for parallel processing
            tasks = []
            for file in files:
                for column in columns:
                    # Use smart extraction method
                    task = self.extract_cell_value_smart(file, column)
                    tasks.append(task)
            
            # Execute tasks in parallel
            logger.info(f"Starting parallel extraction: {len(tasks)} tasks")
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            # Save results to database
            saved_count = 0
            error_count = 0
            
            for result in results:
                if isinstance(result, Exception):
                    error_count += 1
                    logger.error(f"Extraction task failed: {result}")
                    continue
                
                if result.get("error"):
                    error_count += 1
                    continue
                
                # Check if cell already exists
                existing_cell = self.db.query(TabularCell).filter(
                    and_(
                        TabularCell.file_id == result["file_id"],
                        TabularCell.column_id == result["column_id"]
                    )
                ).first()
                
                if existing_cell:
                    # Update existing cell with all new fields
                    existing_cell.cell_value = result.get("cell_value")
                    existing_cell.verbatim_extract = result.get("verbatim_extract")
                    existing_cell.reasoning = result.get("reasoning")
                    existing_cell.source_references = result.get("source_references")
                    existing_cell.confidence_score = result.get("confidence_score")
                    # Extract source_page from first source_reference if available
                    source_refs = result.get("source_references", [])
                    if source_refs and isinstance(source_refs, list) and len(source_refs) > 0:
                        first_ref = source_refs[0]
                        if isinstance(first_ref, dict):
                            existing_cell.source_page = first_ref.get("page")
                            existing_cell.source_section = first_ref.get("section")
                    else:
                        existing_cell.source_page = result.get("source_page")
                        existing_cell.source_section = result.get("source_section")
                        existing_cell.status = "completed"
                        existing_cell.updated_at = datetime.utcnow()
                else:
                    # Extract source_page from first source_reference if available
                    source_refs = result.get("source_references", [])
                    source_page = result.get("source_page")
                    source_section = result.get("source_section")
                    if source_refs and isinstance(source_refs, list) and len(source_refs) > 0:
                        first_ref = source_refs[0]
                        if isinstance(first_ref, dict):
                            source_page = first_ref.get("page") or source_page
                            source_section = first_ref.get("section") or source_section
                    
                    # Create new cell with all fields
                    cell = TabularCell(
                        tabular_review_id=review_id,
                        file_id=result["file_id"],
                        column_id=result["column_id"],
                        cell_value=result.get("cell_value"),
                        verbatim_extract=result.get("verbatim_extract"),
                        reasoning=result.get("reasoning"),
                        source_references=result.get("source_references"),
                        confidence_score=result.get("confidence_score"),
                        source_page=source_page,
                        source_section=source_section,
                        status="completed"
                    )
                    self.db.add(cell)
                
                saved_count += 1
            
            self.db.commit()
            
            # Update review status
            review.status = "completed"
            review.updated_at = datetime.utcnow()
            self.db.commit()
            
            logger.info(f"Extraction completed: {saved_count} cells saved, {error_count} errors")
            
            return {
                "status": "completed",
                "saved_count": saved_count,
                "error_count": error_count,
                "total_tasks": len(tasks)
            }
            
        except Exception as e:
            logger.error(f"Error during extraction: {e}", exc_info=True)
            review.status = "draft"
            self.db.commit()
            raise
    
    async def run_column_extraction(self, review_id: str, column_id: str, user_id: str) -> Dict[str, Any]:
        """Run extraction for a specific column across all documents"""
        # Verify review belongs to user
        review = self.db.query(TabularReview).filter(
            and_(TabularReview.id == review_id, TabularReview.user_id == user_id)
        ).first()
        
        if not review:
            raise ValueError(f"Tabular review {review_id} not found or access denied")
        
        # Get the specific column
        column = self.db.query(TabularColumn).filter(
            and_(
                TabularColumn.id == column_id,
                TabularColumn.tabular_review_id == review_id
            )
        ).first()
        
        if not column:
            raise ValueError(f"Column {column_id} not found in review {review_id}")
        
        # Get files - filter by selected_file_ids if specified
        files_query = self.db.query(File).filter(File.case_id == review.case_id)
        if review.selected_file_ids:
            files_query = files_query.filter(File.id.in_(review.selected_file_ids))
        files = files_query.all()
        
        if not files:
            raise ValueError("No files found for this case")
        
        # Create tasks for parallel processing (only for this column)
        tasks = []
        for file in files:
            task = self.extract_cell_value_smart(file, column)
            tasks.append(task)
        
        # Execute tasks in parallel
        logger.info(f"Starting parallel extraction for column {column_id}: {len(tasks)} tasks")
        results = await asyncio.gather(*tasks, return_exceptions=True)
        
        # Save results to database
        saved_count = 0
        error_count = 0
        
        for result in results:
            if isinstance(result, Exception):
                error_count += 1
                logger.error(f"Extraction task failed: {result}")
                continue
            
            if result.get("error"):
                error_count += 1
                continue
            
            # Check if cell already exists
            existing_cell = self.db.query(TabularCell).filter(
                and_(
                    TabularCell.file_id == result["file_id"],
                    TabularCell.column_id == result["column_id"]
                )
            ).first()
            
            if existing_cell:
                # Update existing cell with all new fields
                existing_cell.cell_value = result.get("cell_value")
                existing_cell.verbatim_extract = result.get("verbatim_extract")
                existing_cell.reasoning = result.get("reasoning")
                existing_cell.source_references = result.get("source_references")
                existing_cell.confidence_score = result.get("confidence_score")
                # Extract source_page from first source_reference if available
                source_refs = result.get("source_references", [])
                if source_refs and isinstance(source_refs, list) and len(source_refs) > 0:
                    first_ref = source_refs[0]
                    if isinstance(first_ref, dict):
                        existing_cell.source_page = first_ref.get("page")
                        existing_cell.source_section = first_ref.get("section")
                else:
                    existing_cell.source_page = result.get("source_page")
                    existing_cell.source_section = result.get("source_section")
                existing_cell.status = "completed"
                existing_cell.updated_at = datetime.utcnow()
            else:
                # Extract source_page from first source_reference if available
                source_refs = result.get("source_references", [])
                source_page = result.get("source_page")
                source_section = result.get("source_section")
                if source_refs and isinstance(source_refs, list) and len(source_refs) > 0:
                    first_ref = source_refs[0]
                    if isinstance(first_ref, dict):
                        source_page = first_ref.get("page") or source_page
                        source_section = first_ref.get("section") or source_section
                
                # Create new cell with all fields
                cell = TabularCell(
                    tabular_review_id=review_id,
                    file_id=result["file_id"],
                    column_id=result["column_id"],
                    cell_value=result.get("cell_value"),
                    verbatim_extract=result.get("verbatim_extract"),
                    reasoning=result.get("reasoning"),
                    source_references=result.get("source_references"),
                    confidence_score=result.get("confidence_score"),
                    source_page=source_page,
                    source_section=source_section,
                    status="completed"
                )
                self.db.add(cell)
            
            saved_count += 1
        
        self.db.commit()
        
        logger.info(f"Column extraction completed: {saved_count} cells saved, {error_count} errors")
        
        return {
            "status": "completed",
            "saved_count": saved_count,
            "error_count": error_count,
            "total_tasks": len(tasks),
            "column_id": column_id
        }
    
    def update_selected_files(
        self,
        review_id: str,
        file_ids: List[str],
        user_id: str
    ) -> TabularReview:
        """Update selected files for a tabular review"""
        # Verify review belongs to user
        review = self.db.query(TabularReview).filter(
            and_(TabularReview.id == review_id, TabularReview.user_id == user_id)
        ).first()
        
        if not review:
            raise ValueError(f"Tabular review {review_id} not found or access denied")
        
        # Verify files belong to the case
        if file_ids:
            files = self.db.query(File).filter(
                and_(
                    File.id.in_(file_ids),
                    File.case_id == review.case_id
                )
            ).all()
            if len(files) != len(file_ids):
                raise ValueError("Some selected files do not belong to this case")
        
        review.selected_file_ids = file_ids if file_ids else None
        review.updated_at = datetime.utcnow()
        self.db.commit()
        self.db.refresh(review)
        
        logger.info(f"Updated selected files for review {review_id}: {len(file_ids) if file_ids else 0} files")
        return review
    
    def mark_as_reviewed(
        self,
        review_id: str,
        file_id: str,
        user_id: str,
        status: str
    ) -> TabularDocumentStatus:
        """Mark document as reviewed"""
        # Verify review belongs to user
        review = self.db.query(TabularReview).filter(
            and_(TabularReview.id == review_id, TabularReview.user_id == user_id)
        ).first()
        
        if not review:
            raise ValueError(f"Tabular review {review_id} not found or access denied")
        
        # Get or create status
        doc_status = self.db.query(TabularDocumentStatus).filter(
            and_(
                TabularDocumentStatus.tabular_review_id == review_id,
                TabularDocumentStatus.file_id == file_id,
                TabularDocumentStatus.user_id == user_id
            )
        ).first()
        
        if not doc_status:
            doc_status = TabularDocumentStatus(
                tabular_review_id=review_id,
                file_id=file_id,
                user_id=user_id,
                status=status
            )
            self.db.add(doc_status)
        else:
            doc_status.status = status
            doc_status.updated_at = datetime.utcnow()
        
        self.db.commit()
        self.db.refresh(doc_status)
        
        return doc_status
    
    def update_cell(
        self,
        review_id: str,
        file_id: str,
        column_id: str,
        cell_value: str,
        user_id: str,
        is_manual_override: bool = True
    ) -> TabularCell:
        """Update a cell value (manual edit) by file_id and column_id"""
        # Verify review belongs to user
        review = self.db.query(TabularReview).filter(
            and_(TabularReview.id == review_id, TabularReview.user_id == user_id)
        ).first()
        
        if not review:
            raise ValueError(f"Tabular review {review_id} not found or access denied")
        
        # Get or create the cell (using unique constraint on file_id + column_id)
        cell = self.db.query(TabularCell).filter(
            and_(
                TabularCell.tabular_review_id == review_id,
                TabularCell.file_id == file_id,
                TabularCell.column_id == column_id
            )
        ).first()
        
        # Save previous value for history
        previous_cell_value = cell.cell_value if cell else None
        change_type = "updated" if cell else "created"
        
        if not cell:
            # Create new cell if it doesn't exist
            cell = TabularCell(
                tabular_review_id=review_id,
                file_id=file_id,
                column_id=column_id,
                cell_value=cell_value,
                status="reviewed" if is_manual_override else "pending"
            )
            self.db.add(cell)
            self.db.flush()  # Flush to get the cell ID
        else:
            # Check if cell is locked by another user
            if cell.locked_by and cell.locked_by != user_id:
                # Check if lock has expired
                if cell.lock_expires_at and cell.lock_expires_at > datetime.utcnow():
                    locked_user = self.db.query(User).filter(User.id == cell.locked_by).first()
                    locked_user_name = locked_user.full_name if locked_user else "Unknown"
                    raise ValueError(f"Cell is locked by {locked_user_name}")
                # Lock has expired, allow update and unlock
                cell.locked_by = None
                cell.locked_at = None
                cell.lock_expires_at = None
            
            # Update existing cell
            cell.cell_value = cell_value
            cell.status = "reviewed" if is_manual_override else cell.status
            cell.updated_at = datetime.utcnow()
            # For manual edits, clear AI-related fields
            if is_manual_override:
                cell.verbatim_extract = None
                cell.reasoning = None
                cell.source_references = None
                cell.confidence_score = None
                cell.source_page = None
                cell.source_section = None
            
            # Unlock cell after update (if it was locked by this user)
            if cell.locked_by == user_id:
                cell.locked_by = None
                cell.locked_at = None
                cell.lock_expires_at = None
        
        self.db.commit()
        self.db.refresh(cell)
        
        # Log the change to history
        try:
            from app.services.cell_history_service import CellHistoryService
            history_service = CellHistoryService(self.db)
            history_service.log_cell_change(
                cell=cell,
                change_type=change_type,
                changed_by=user_id,
                previous_cell_value=previous_cell_value,
                change_reason="Manual edit" if is_manual_override else "Cell update"
            )
        except Exception as e:
            logger.warning(f"Failed to log cell history: {e}", exc_info=True)
            # Don't fail the update if history logging fails
        
        logger.info(f"Updated cell for file {file_id}, column {column_id} in review {review_id}")
        return cell
    
    def bulk_update_status(
        self,
        review_id: str,
        file_ids: List[str],
        status: str,
        user_id: str
    ) -> int:
        """Bulk update document status for multiple files"""
        # Verify review belongs to user
        review = self.db.query(TabularReview).filter(
            and_(TabularReview.id == review_id, TabularReview.user_id == user_id)
        ).first()
        
        if not review:
            raise ValueError(f"Tabular review {review_id} not found or access denied")
        
        # Verify files belong to the case
        files = self.db.query(File).filter(
            and_(
                File.id.in_(file_ids),
                File.case_id == review.case_id
            )
        ).all()
        
        if len(files) != len(file_ids):
            raise ValueError("Some files do not belong to this case")
        
        updated_count = 0
        for file_id in file_ids:
            # Get or create document status
            doc_status = self.db.query(TabularDocumentStatus).filter(
                and_(
                    TabularDocumentStatus.tabular_review_id == review_id,
                    TabularDocumentStatus.file_id == file_id,
                    TabularDocumentStatus.user_id == user_id
                )
            ).first()
            
            if doc_status:
                doc_status.status = status
                doc_status.updated_at = datetime.utcnow()
            else:
                doc_status = TabularDocumentStatus(
                    tabular_review_id=review_id,
                    file_id=file_id,
                    user_id=user_id,
                    status=status
                )
                self.db.add(doc_status)
            
            updated_count += 1
        
        self.db.commit()
        logger.info(f"Bulk updated status for {updated_count} files in review {review_id}")
        return updated_count
    
    async def bulk_run_extraction(
        self,
        review_id: str,
        file_ids: List[str],
        column_ids: List[str],
        user_id: str
    ) -> Dict[str, Any]:
        """Bulk run extraction for specific files and columns"""
        # Verify review belongs to user
        review = self.db.query(TabularReview).filter(
            and_(TabularReview.id == review_id, TabularReview.user_id == user_id)
        ).first()
        
        if not review:
            raise ValueError(f"Tabular review {review_id} not found or access denied")
        
        # Get files
        files = self.db.query(File).filter(
            and_(
                File.id.in_(file_ids),
                File.case_id == review.case_id
            )
        ).all()
        
        if len(files) != len(file_ids):
            raise ValueError("Some files do not belong to this case")
        
        # Get columns
        columns = self.db.query(TabularColumn).filter(
            and_(
                TabularColumn.tabular_review_id == review_id,
                TabularColumn.id.in_(column_ids)
            )
        ).all()
        
        if len(columns) != len(column_ids):
            raise ValueError("Some columns do not belong to this review")
        
        # Create tasks for parallel processing
        tasks = []
        for file in files:
            for column in columns:
                task = self.extract_cell_value_smart(file, column)
                tasks.append(task)
        
        # Execute tasks in parallel
        logger.info(f"Starting bulk extraction: {len(tasks)} tasks for {len(file_ids)} files, {len(column_ids)} columns")
        results = await asyncio.gather(*tasks, return_exceptions=True)
        
        # Save results to database
        saved_count = 0
        error_count = 0
        
        for result in results:
            if isinstance(result, Exception):
                error_count += 1
                logger.error(f"Extraction task failed: {result}")
                continue
            
            if result.get("error"):
                error_count += 1
                continue
            
            # Check if cell already exists
            existing_cell = self.db.query(TabularCell).filter(
                and_(
                    TabularCell.file_id == result["file_id"],
                    TabularCell.column_id == result["column_id"]
                )
            ).first()
            
            if existing_cell:
                # Update existing cell
                existing_cell.cell_value = result.get("cell_value")
                existing_cell.verbatim_extract = result.get("verbatim_extract")
                existing_cell.reasoning = result.get("reasoning")
                existing_cell.source_references = result.get("source_references")
                existing_cell.confidence_score = result.get("confidence_score")
                source_refs = result.get("source_references", [])
                if source_refs and isinstance(source_refs, list) and len(source_refs) > 0:
                    first_ref = source_refs[0]
                    if isinstance(first_ref, dict):
                        existing_cell.source_page = first_ref.get("page")
                        existing_cell.source_section = first_ref.get("section")
                else:
                    existing_cell.source_page = result.get("source_page")
                    existing_cell.source_section = result.get("source_section")
                existing_cell.status = "completed"
                existing_cell.updated_at = datetime.utcnow()
            else:
                # Create new cell
                source_refs = result.get("source_references", [])
                source_page = result.get("source_page")
                source_section = result.get("source_section")
                if source_refs and isinstance(source_refs, list) and len(source_refs) > 0:
                    first_ref = source_refs[0]
                    if isinstance(first_ref, dict):
                        source_page = first_ref.get("page") or source_page
                        source_section = first_ref.get("section") or source_section
                
                cell = TabularCell(
                    tabular_review_id=review_id,
                    file_id=result["file_id"],
                    column_id=result["column_id"],
                    cell_value=result.get("cell_value"),
                    verbatim_extract=result.get("verbatim_extract"),
                    reasoning=result.get("reasoning"),
                    source_references=result.get("source_references"),
                    confidence_score=result.get("confidence_score"),
                    source_page=source_page,
                    source_section=source_section,
                    status="completed"
                )
                self.db.add(cell)
            
            saved_count += 1
        
        self.db.commit()
        
        logger.info(f"Bulk extraction completed: {saved_count} cells saved, {error_count} errors")
        
        return {
            "status": "completed",
            "saved_count": saved_count,
            "error_count": error_count,
            "total_tasks": len(tasks),
            "files_processed": len(file_ids),
            "columns_processed": len(column_ids)
        }
    
    def bulk_delete_rows(
        self,
        review_id: str,
        file_ids: List[str],
        user_id: str
    ) -> int:
        """Bulk delete rows (remove files from selected_file_ids)"""
        # Verify review belongs to user
        review = self.db.query(TabularReview).filter(
            and_(TabularReview.id == review_id, TabularReview.user_id == user_id)
        ).first()
        
        if not review:
            raise ValueError(f"Tabular review {review_id} not found or access denied")
        
        # Update selected_file_ids to remove deleted files
        if review.selected_file_ids:
            updated_file_ids = [fid for fid in review.selected_file_ids if fid not in file_ids]
            review.selected_file_ids = updated_file_ids if updated_file_ids else None
        else:
            # If no selected_file_ids, all files are included, so we can't "delete" them
            # Instead, we'll set selected_file_ids to exclude these files
            # Get all files for the case
            all_files = self.db.query(File).filter(File.case_id == review.case_id).all()
            all_file_ids = [f.id for f in all_files]
            updated_file_ids = [fid for fid in all_file_ids if fid not in file_ids]
            review.selected_file_ids = updated_file_ids if updated_file_ids else None
        
        # Delete cells for these files (cascade should handle this, but explicit for clarity)
        deleted_cells = self.db.query(TabularCell).filter(
            and_(
                TabularCell.tabular_review_id == review_id,
                TabularCell.file_id.in_(file_ids)
            )
        ).delete(synchronize_session=False)
        
        # Delete document statuses for these files
        deleted_statuses = self.db.query(TabularDocumentStatus).filter(
            and_(
                TabularDocumentStatus.tabular_review_id == review_id,
                TabularDocumentStatus.file_id.in_(file_ids)
            )
        ).delete(synchronize_session=False)
        
        review.updated_at = datetime.utcnow()
        self.db.commit()
        
        logger.info(f"Bulk deleted {len(file_ids)} rows from review {review_id} (removed {deleted_cells} cells, {deleted_statuses} statuses)")
        return len(file_ids)
    
    def export_to_csv(self, review_id: str, user_id: str) -> str:
        """Export tabular review to CSV format"""
        import csv
        import io
        
        data = self.get_table_data(review_id, user_id)
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        header = ["Document"] + [col["column_label"] for col in data["columns"]]
        writer.writerow(header)
        
        # Write rows
        for row in data["rows"]:
            csv_row = [row["file_name"]]
            for col in data["columns"]:
                cell = row["cells"].get(col["id"], {})
                csv_row.append(cell.get("cell_value") or "")
            writer.writerow(csv_row)
        
        return output.getvalue()
    
    def export_to_excel(self, review_id: str, user_id: str) -> bytes:
        """Export tabular review to Excel format"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            import io
            
            data = self.get_table_data(review_id, user_id)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Tabular Review"
            
            # Write header
            header = ["Document"] + [col["column_label"] for col in data["columns"]]
            for col_num, value in enumerate(header, start=1):
                cell = ws.cell(row=1, column=col_num, value=value)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Write rows
            for row_num, row in enumerate(data["rows"], start=2):
                ws.cell(row=row_num, column=1, value=row["file_name"])
                for col_num, col in enumerate(data["columns"], start=2):
                    cell = row["cells"].get(col["id"], {})
                    ws.cell(row=row_num, column=col_num, value=cell.get("cell_value") or "")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()
            
        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            # Fallback to CSV if openpyxl not available
            csv_content = self.export_to_csv(review_id, user_id)
            return csv_content.encode('utf-8')
    
    def create_timeline_table_from_results(
        self,
        case_id: str,
        user_id: str,
        name: str = "Хронология событий"
    ) -> TabularReview:
        """Создает таблицу из результатов Timeline агента"""
        from app.models.analysis import TimelineEvent
        
        # Get timeline events
        events = self.db.query(TimelineEvent).filter(
            TimelineEvent.case_id == case_id
        ).order_by(TimelineEvent.date).all()
        
        if not events:
            raise ValueError("No timeline events found for this case")
        
        # Create review
        review = self.create_tabular_review(
            case_id=case_id,
            user_id=user_id,
            name=name,
            description="Автоматически созданная таблица из хронологии событий"
        )
        
        # Add columns
        self.add_column(
            review_id=review.id,
            column_label="Дата",
            column_type="date",
            prompt="Дата события",
            user_id=user_id
        )
        
        self.add_column(
            review_id=review.id,
            column_label="Тип события",
            column_type="text",
            prompt="Тип события",
            user_id=user_id
        )
        
        self.add_column(
            review_id=review.id,
            column_label="Описание",
            column_type="text",
            prompt="Описание события",
            user_id=user_id
        )
        
        self.add_column(
            review_id=review.id,
            column_label="Источник",
            column_type="text",
            prompt="Документ-источник",
            user_id=user_id
        )
        
        # Get files for events
        from app.models.case import File
        files_dict = {}
        for event in events:
            source_doc = event.source_document
            if source_doc and source_doc not in files_dict:
                file = self.db.query(File).filter(
                    File.case_id == case_id,
                    File.filename == source_doc
                ).first()
                if file:
                    files_dict[source_doc] = file
        
        # Create cells from events
        for event in events:
            file = files_dict.get(event.source_document)
            if not file:
                continue
            
            # Date cell
            cell_date = TabularCell(
                tabular_review_id=review.id,
                file_id=file.id,
                column_id=review.columns[0].id,
                cell_value=event.date.strftime("%Y-%m-%d") if event.date else "",
                confidence_score=0.9
            )
            self.db.add(cell_date)
            
            # Event type cell
            cell_type = TabularCell(
                tabular_review_id=review.id,
                file_id=file.id,
                column_id=review.columns[1].id,
                cell_value=event.event_type or "",
                confidence_score=0.9
            )
            self.db.add(cell_type)
            
            # Description cell
            cell_desc = TabularCell(
                tabular_review_id=review.id,
                file_id=file.id,
                column_id=review.columns[2].id,
                cell_value=event.description,
                confidence_score=0.9
            )
            self.db.add(cell_desc)
            
            # Source cell
            cell_source = TabularCell(
                tabular_review_id=review.id,
                file_id=file.id,
                column_id=review.columns[3].id,
                cell_value=event.source_document,
                confidence_score=1.0
            )
            self.db.add(cell_source)
        
        self.db.commit()
        logger.info(f"Created timeline table {review.id} with {len(events)} events")
        return review
    
    def create_entities_table_from_results(
        self,
        case_id: str,
        user_id: str,
        name: str = "Извлеченные сущности"
    ) -> TabularReview:
        """Создает таблицу из результатов Entity Extraction агента"""
        from app.models.analysis import ExtractedEntity
        
        # Get entities
        entities = self.db.query(ExtractedEntity).filter(
            ExtractedEntity.case_id == case_id
        ).all()
        
        if not entities:
            raise ValueError("No entities found for this case")
        
        # Create review
        review = self.create_tabular_review(
            case_id=case_id,
            user_id=user_id,
            name=name,
            description="Автоматически созданная таблица из извлеченных сущностей"
        )
        
        # Add columns
        self.add_column(
            review_id=review.id,
            column_label="Сущность",
            column_type="text",
            prompt="Название сущности",
            user_id=user_id
        )
        
        self.add_column(
            review_id=review.id,
            column_label="Тип",
            column_type="text",
            prompt="Тип сущности (PERSON, ORG, etc.)",
            user_id=user_id
        )
        
        self.add_column(
            review_id=review.id,
            column_label="Контекст",
            column_type="text",
            prompt="Контекст, в котором найдена сущность",
            user_id=user_id
        )
        
        self.add_column(
            review_id=review.id,
            column_label="Источник",
            column_type="text",
            prompt="Документ-источник",
            user_id=user_id
        )
        
        # Get files for entities
        from app.models.case import File
        files_dict = {}
        for entity in entities:
            if entity.file_id and entity.file_id not in files_dict:
                file = self.db.query(File).filter(File.id == entity.file_id).first()
                if file:
                    files_dict[entity.file_id] = file
        
        # Create cells from entities
        for entity in entities:
            file = files_dict.get(entity.file_id) if entity.file_id else None
            if not file:
                continue
            
            # Entity text cell
            cell_entity = TabularCell(
                tabular_review_id=review.id,
                file_id=file.id,
                column_id=review.columns[0].id,
                cell_value=entity.entity_text,
                confidence_score=float(entity.confidence) if entity.confidence else 0.8
            )
            self.db.add(cell_entity)
            
            # Entity type cell
            cell_type = TabularCell(
                tabular_review_id=review.id,
                file_id=file.id,
                column_id=review.columns[1].id,
                cell_value=entity.entity_type,
                confidence_score=0.9
            )
            self.db.add(cell_type)
            
            # Context cell
            cell_context = TabularCell(
                tabular_review_id=review.id,
                file_id=file.id,
                column_id=review.columns[2].id,
                cell_value=entity.context or "",
                confidence_score=0.8
            )
            self.db.add(cell_context)
            
            # Source cell
            cell_source = TabularCell(
                tabular_review_id=review.id,
                file_id=file.id,
                column_id=review.columns[3].id,
                cell_value=entity.source_document or file.name,
                confidence_score=1.0
            )
            self.db.add(cell_source)
        
        self.db.commit()
        logger.info(f"Created entities table {review.id} with {len(entities)} entities")
        return review
    
    def create_discrepancies_table_from_results(
        self,
        case_id: str,
        user_id: str,
        name: str = "Найденные противоречия"
    ) -> TabularReview:
        """Создает таблицу из результатов Discrepancy агента"""
        from app.models.analysis import Discrepancy
        
        # Get discrepancies
        discrepancies = self.db.query(Discrepancy).filter(
            Discrepancy.case_id == case_id
        ).all()
        
        if not discrepancies:
            raise ValueError("No discrepancies found for this case")
        
        # Create review
        review = self.create_tabular_review(
            case_id=case_id,
            user_id=user_id,
            name=name,
            description="Автоматически созданная таблица из найденных противоречий"
        )
        
        # Add columns
        self.add_column(
            review_id=review.id,
            column_label="Тип противоречия",
            column_type="text",
            prompt="Тип противоречия",
            user_id=user_id
        )
        
        self.add_column(
            review_id=review.id,
            column_label="Серьезность",
            column_type="text",
            prompt="Уровень серьезности (HIGH, MEDIUM, LOW)",
            user_id=user_id
        )
        
        self.add_column(
            review_id=review.id,
            column_label="Описание",
            column_type="text",
            prompt="Описание противоречия",
            user_id=user_id
        )
        
        self.add_column(
            review_id=review.id,
            column_label="Документы",
            column_type="text",
            prompt="Документы с противоречиями",
            user_id=user_id
        )
        
        # Get files
        from app.models.case import File
        files = self.db.query(File).filter(File.case_id == case_id).all()
        if not files:
            raise ValueError("No files found for this case")
        
        # Create cells from discrepancies
        # Check for existing cells to avoid duplicates
        from app.models.tabular_review import TabularCell
        
        for disc in discrepancies:
            # Use first file as placeholder (discrepancies don't have direct file_id)
            file = files[0]
            
            # Helper function to check and create cell
            def get_or_create_cell(column_idx, value, confidence=0.9):
                column_id = review.columns[column_idx].id
                # Check if cell already exists
                existing = self.db.query(TabularCell).filter(
                    TabularCell.tabular_review_id == review.id,
                    TabularCell.file_id == file.id,
                    TabularCell.column_id == column_id
                ).first()
                
                if existing:
                    # Update existing cell
                    existing.cell_value = value
                    existing.confidence_score = confidence
                    return existing
                else:
                    # Create new cell
                    return TabularCell(
                        tabular_review_id=review.id,
                        file_id=file.id,
                        column_id=column_id,
                        cell_value=value,
                        confidence_score=confidence
                    )
            
            # Type cell
            cell_type = get_or_create_cell(0, disc.type, 0.9)
            if not self.db.is_modified(cell_type, include_collections=False) and cell_type not in self.db.new:
                self.db.add(cell_type)
            
            # Severity cell
            cell_severity = get_or_create_cell(1, disc.severity, 0.9)
            if not self.db.is_modified(cell_severity, include_collections=False) and cell_severity not in self.db.new:
                self.db.add(cell_severity)
            
            # Description cell
            cell_desc = get_or_create_cell(2, disc.description, 0.9)
            if not self.db.is_modified(cell_desc, include_collections=False) and cell_desc not in self.db.new:
                self.db.add(cell_desc)
            
            # Documents cell
            source_docs = disc.source_documents
            if isinstance(source_docs, list):
                docs_str = ", ".join(source_docs)
            else:
                docs_str = str(source_docs)
            
            cell_docs = get_or_create_cell(3, docs_str, 1.0)
            if not self.db.is_modified(cell_docs, include_collections=False) and cell_docs not in self.db.new:
                self.db.add(cell_docs)
        
        try:
            self.db.commit()
        except Exception as e:
            self.db.rollback()
            raise
        logger.info(f"Created discrepancies table {review.id} with {len(discrepancies)} discrepancies")
        return review
    
    def create_risks_table_from_results(
        self,
        case_id: str,
        user_id: str,
        name: str = "Выявленные риски"
    ) -> TabularReview:
        """Создает таблицу из результатов Risk агента"""
        from app.models.analysis import Risk
        
        # Get risks
        risks = self.db.query(Risk).filter(
            Risk.case_id == case_id
        ).all()
        
        if not risks:
            raise ValueError("No risks found for this case")
        
        # Create review
        review = self.create_tabular_review(
            case_id=case_id,
            user_id=user_id,
            name=name,
            description="Автоматически созданная таблица из выявленных рисков"
        )
        
        # Add columns
        self.add_column(
            review_id=review.id,
            column_label="Название риска",
            column_type="text",
            prompt="Название риска",
            user_id=user_id
        )
        
        self.add_column(
            review_id=review.id,
            column_label="Категория",
            column_type="text",
            prompt="Категория риска",
            user_id=user_id
        )
        
        self.add_column(
            review_id=review.id,
            column_label="Вероятность",
            column_type="text",
            prompt="Вероятность (HIGH, MEDIUM, LOW)",
            user_id=user_id
        )
        
        self.add_column(
            review_id=review.id,
            column_label="Влияние",
            column_type="text",
            prompt="Влияние (HIGH, MEDIUM, LOW)",
            user_id=user_id
        )
        
        self.add_column(
            review_id=review.id,
            column_label="Описание",
            column_type="text",
            prompt="Описание риска",
            user_id=user_id
        )
        
        # Get files
        from app.models.case import File
        files = self.db.query(File).filter(File.case_id == case_id).all()
        if not files:
            raise ValueError("No files found for this case")
        
        # Create cells from risks
        for risk in risks:
            # Use first file as placeholder
            file = files[0]
            
            # Name cell
            cell_name = TabularCell(
                tabular_review_id=review.id,
                file_id=file.id,
                column_id=review.columns[0].id,
                cell_value=risk.risk_name,
                confidence_score=0.9
            )
            self.db.add(cell_name)
            
            # Category cell
            cell_category = TabularCell(
                tabular_review_id=review.id,
                file_id=file.id,
                column_id=review.columns[1].id,
                cell_value=risk.risk_category,
                confidence_score=0.9
            )
            self.db.add(cell_category)
            
            # Probability cell
            cell_prob = TabularCell(
                tabular_review_id=review.id,
                file_id=file.id,
                column_id=review.columns[2].id,
                cell_value=risk.probability,
                confidence_score=0.9
            )
            self.db.add(cell_prob)
            
            # Impact cell
            cell_impact = TabularCell(
                tabular_review_id=review.id,
                file_id=file.id,
                column_id=review.columns[3].id,
                cell_value=risk.impact,
                confidence_score=0.9
            )
            self.db.add(cell_impact)
            
            # Description cell
            cell_desc = TabularCell(
                tabular_review_id=review.id,
                file_id=file.id,
                column_id=review.columns[4].id,
                cell_value=risk.description,
                confidence_score=0.9
            )
            self.db.add(cell_desc)
        
        self.db.commit()
        logger.info(f"Created risks table {review.id} with {len(risks)} risks")
        return review

    def create_key_facts_table_from_results(
        self,
        case_id: str,
        user_id: str,
        name: str = "Ключевые факты"
    ) -> TabularReview:
        """Создает таблицу из результатов Key Facts агента"""
        from app.models.analysis import AnalysisResult
        
        # Get key_facts from AnalysisResult
        key_facts_result = self.db.query(AnalysisResult).filter(
            AnalysisResult.case_id == case_id,
            AnalysisResult.analysis_type == "key_facts"
        ).order_by(AnalysisResult.created_at.desc()).first()
        
        if not key_facts_result or not key_facts_result.result_data:
            raise ValueError("No key_facts found for this case")
        
        result_data = key_facts_result.result_data
        
        # Create review
        review = self.create_tabular_review(
            case_id=case_id,
            user_id=user_id,
            name=name,
            description="Автоматически созданная таблица из ключевых фактов"
        )
        
        # Add columns based on key_facts structure
        # Обычно key_facts содержит: parties, amounts, dates, key_terms, etc.
        self.add_column(
            review_id=review.id,
            column_label="Категория",
            column_type="text",
            prompt="Категория факта",
            user_id=user_id
        )
        
        self.add_column(
            review_id=review.id,
            column_label="Значение",
            column_type="text",
            prompt="Значение факта",
            user_id=user_id
        )
        
        self.add_column(
            review_id=review.id,
            column_label="Описание",
            column_type="text",
            prompt="Описание факта",
            user_id=user_id
        )
        
        # Get files
        from app.models.case import File
        files = self.db.query(File).filter(File.case_id == case_id).all()
        if not files:
            raise ValueError("No files found for this case")
        
        # Use first file as placeholder (key_facts are case-level, not file-level)
        file = files[0]
        
        # Extract key facts from result_data
        facts_to_add = []
        
        # Parties
        if result_data.get("parties"):
            parties = result_data["parties"]
            if isinstance(parties, dict):
                for role, name in parties.items():
                    facts_to_add.append({
                        "category": f"Сторона: {role}",
                        "value": str(name),
                        "description": f"Сторона дела: {role}"
                    })
        
        # Amounts
        if result_data.get("amounts"):
            amounts = result_data["amounts"]
            if isinstance(amounts, dict):
                for key, value in amounts.items():
                    facts_to_add.append({
                        "category": f"Сумма: {key}",
                        "value": str(value),
                        "description": f"Финансовая информация: {key}"
                    })
        
        # Dates
        if result_data.get("dates"):
            dates = result_data["dates"]
            if isinstance(dates, dict):
                for key, value in dates.items():
                    facts_to_add.append({
                        "category": f"Дата: {key}",
                        "value": str(value),
                        "description": f"Важная дата: {key}"
                    })
        
        # Key terms
        if result_data.get("key_terms"):
            key_terms = result_data["key_terms"]
            if isinstance(key_terms, list):
                for term in key_terms:
                    facts_to_add.append({
                        "category": "Ключевой термин",
                        "value": str(term),
                        "description": "Важный термин из документов"
                    })
        
        # If no structured data, add summary
        if not facts_to_add and result_data.get("summary"):
            facts_to_add.append({
                "category": "Резюме",
                "value": result_data["summary"][:500],
                "description": "Общее резюме ключевых фактов"
            })
        
        # Create cells from facts
        for fact in facts_to_add:
            # Category cell
            cell_category = TabularCell(
                tabular_review_id=review.id,
                file_id=file.id,
                column_id=review.columns[0].id,
                cell_value=fact["category"],
                confidence_score=0.9
            )
            self.db.add(cell_category)
            
            # Value cell
            cell_value = TabularCell(
                tabular_review_id=review.id,
                file_id=file.id,
                column_id=review.columns[1].id,
                cell_value=fact["value"],
                confidence_score=0.9
            )
            self.db.add(cell_value)
            
            # Description cell
            cell_desc = TabularCell(
                tabular_review_id=review.id,
                file_id=file.id,
                column_id=review.columns[2].id,
                cell_value=fact["description"],
                confidence_score=0.9
            )
            self.db.add(cell_desc)
        
        self.db.commit()
        logger.info(f"Created key_facts table {review.id} with {len(facts_to_add)} facts")
        return review
    
    def lock_cell(
        self,
        review_id: str,
        file_id: str,
        column_id: str,
        user_id: str,
        lock_duration_seconds: int = 300  # Default 5 minutes
    ) -> TabularCell:
        """Lock a cell for editing by a specific user"""
        # Verify review belongs to user
        review = self.db.query(TabularReview).filter(
            and_(TabularReview.id == review_id, TabularReview.user_id == user_id)
        ).first()
        
        if not review:
            raise ValueError(f"Tabular review {review_id} not found or access denied")
        
        # Get or create the cell
        cell = self.db.query(TabularCell).filter(
            and_(
                TabularCell.tabular_review_id == review_id,
                TabularCell.file_id == file_id,
                TabularCell.column_id == column_id
            )
        ).first()
        
        if not cell:
            # Create cell if it doesn't exist
            cell = TabularCell(
                tabular_review_id=review_id,
                file_id=file_id,
                column_id=column_id,
                cell_value=None,
                status="pending"
            )
            self.db.add(cell)
            self.db.flush()
        
        # Check if cell is already locked by another user
        if cell.locked_by and cell.locked_by != user_id:
            # Check if lock has expired
            if cell.lock_expires_at and cell.lock_expires_at > datetime.utcnow():
                locked_user = self.db.query(User).filter(User.id == cell.locked_by).first()
                locked_user_name = locked_user.full_name if locked_user else "Unknown"
                raise ValueError(f"Cell is locked by {locked_user_name}")
            # Lock has expired, allow new lock
        elif cell.locked_by == user_id:
            # Already locked by this user, extend lock
            pass
        
        # Lock the cell
        now = datetime.utcnow()
        cell.locked_by = user_id
        cell.locked_at = now
        cell.lock_expires_at = now + timedelta(seconds=lock_duration_seconds)
        
        self.db.commit()
        self.db.refresh(cell)
        
        logger.info(f"Locked cell for file {file_id}, column {column_id} in review {review_id} by user {user_id}")
        return cell
    
    def unlock_cell(
        self,
        review_id: str,
        file_id: str,
        column_id: str,
        user_id: str
    ) -> TabularCell:
        """Unlock a cell"""
        # Verify review belongs to user
        review = self.db.query(TabularReview).filter(
            and_(TabularReview.id == review_id, TabularReview.user_id == user_id)
        ).first()
        
        if not review:
            raise ValueError(f"Tabular review {review_id} not found or access denied")
        
        # Get the cell
        cell = self.db.query(TabularCell).filter(
            and_(
                TabularCell.tabular_review_id == review_id,
                TabularCell.file_id == file_id,
                TabularCell.column_id == column_id
            )
        ).first()
        
        if not cell:
            raise ValueError(f"Cell not found")
        
        # Check if cell is locked
        if not cell.locked_by:
            # Already unlocked
            return cell
        
        # Check if user can unlock (must be the one who locked it or review owner)
        if cell.locked_by != user_id and review.user_id != user_id:
            raise ValueError("You don't have permission to unlock this cell")
        
        # Unlock the cell
        cell.locked_by = None
        cell.locked_at = None
        cell.lock_expires_at = None
        
        self.db.commit()
        self.db.refresh(cell)
        
        logger.info(f"Unlocked cell for file {file_id}, column {column_id} in review {review_id} by user {user_id}")
        return cell

